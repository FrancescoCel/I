import mysql.connector

def SQLConnect():
    """
    Funzione che crea la connessione a MySQL
    
    Returns
    -------
    conn: connection
        Connessione a MySQL
    """
    conn = mysql.connector.connect(host = "localhost", user = "root",password = "checco")
    #conn = mysql.connector.connect(host = 'localhost', user = 'root', password = 'password')
    #conn = mysql.connector.connect(host = 'localhost', user = 'root', password = 'sole1997')
    
    if checkDb(conn) is True:
        createDb(conn)
        queryUse(conn)
        executeQueries('diagn_title', 'diagn',conn)
        executeQueries('diffsydiw', 'diff',conn)
        executeQueries('symptoms2', 'sym',conn)
        executeQueries('centri','cent',conn)
        executeQueries('centri_cure','cent_cure',conn)
    else:
        print("Database esistente!")
        queryUse(conn)
    
    return conn


def checkDb(conn):
    """
    Funzione che controlla se il database è stato già creato
    
    Parameters
    ----------
    conn: connection
        Connessione a MySQL
    
    Returns
    -------
    boolean
        True se il database esiste, False altrimenti
    """
    cursor = conn.cursor()
    
    queryCheck = """SHOW DATABASES LIKE 'medical';"""
    cursor.execute(queryCheck)
    check = cursor.fetchall()
    
    if not check:
        return True
    else:
        return False


def createDb(conn):
    """
    Funzione che crea il database
    
    Parameters
    ----------
    conn: connection
        Connessione a MySQL
    """
    import pymysql as sql    
    sql.install_as_MySQLdb()
    
    cursor = conn.cursor()
    query1 = """CREATE DATABASE medical; """
    query2 = """USE medical; """
    
    try:
         cursor.execute(query1)
         cursor.execute(query2)
         conn.commit()
    except sql.ProgrammingError:
         pass
    cursor.close()
    

def queryUse(conn):
    """
    Funzione che crea la query che consente di utilizzare il database medical
    
    Parameters
    ----------
    conn: connection
        Connessione a MySQL
    """
    cursor = conn.cursor()
    query = """USE medical;"""
    cursor.execute(query)
    conn.commit()
    
        
def executeQueries(filename,tablename,conn):
    """
    Funzione che crea una tabella ed inserisce i valori al suo interno partendo
    da un file excel
    
    Parameters
    ----------
    filename: str
        Nome del file contenente la tabella
    tablename: str
        Nome da dare alla tabella
    conn: connection
        Connessione al database
    """
    import pymysql as sql    
    sql.install_as_MySQLdb()
    import openpyxl as xl
    
    cursor = conn.cursor()  
    queryDrop = """DROP TABLE IF EXISTS """+ tablename
    
    if filename == 'diagn_title':
        
        #Caricamento del worksheet
        sheet = xl.load_workbook("C:/Users/utente/ICon/Dataset_xlsx/" + filename +".xlsx")
        #sheet = xl.load_workbook("C:/Users/franc/ICon/Dataset_xlsx/" + filename +".xlsx")
        #sheet = xl.load_workbook("C:/Users/nico9/ICon/Dataset_xlsx/" + filename +".xlsx")
        
        table = sheet['id']
    
        #Query per la creazione della tabella del file diagn_title.xlsx
        queryCreate = """CREATE TABLE """+ tablename + """(
                    id VARCHAR(30),
                    title text,
                    cat CHAR(9));"""
 
       #Query per l'inserimento dei dati nella tabella
        insertQuery = """INSERT INTO """+ tablename +"""(
                     id,title,cat) VALUES(LOWER(%s),LOWER(%s),LOWER(%s));"""
        
    elif filename == 'diffsydiw':
        
        #Caricamento del worksheet
        sheet = xl.load_workbook("C:/Users/utente/ICon/Dataset_xlsx/" + filename +".xlsx")
        #sheet = xl.load_workbook("C:/Users/franc/ICon/Dataset_xlsx/" + filename +".xlsx")
        #sheet = xl.load_workbook("C:/Users/nico9/ICon/Dataset_xlsx/" + filename +".xlsx")
        
        table = sheet['syd']
        
        #Query per la creazione della tabella del file diffsydiw.xlsx
        queryCreate = """CREATE TABLE """+ tablename + """(
                    syd CHAR(4),
                    did CHAR(4));"""
        
        #Query per l'inserimento dei dati nella tabella
        insertQuery = """INSERT INTO """+ tablename +"""(
                     syd,did) VALUES(%s,%s);"""
        
    elif filename == 'symptoms2':    
        #Caricamento del worksheet
        sheet = xl.load_workbook("C:/Users/utente/ICon/Dataset_xlsx/" + filename +".xlsx")
        #sheet = xl.load_workbook("C:/Users/franc/ICon/Dataset_xlsx/" + filename +".xlsx")
        #sheet = xl.load_workbook("C:/Users/nico9/ICon/Dataset_xlsx/" + filename +".xlsx")
        
        table = sheet['_id']
        
        #Query per la creazione della tabella del file symptoms2.xlsx
        queryCreate = """CREATE TABLE """+ tablename + """(
                    _id CHAR(4),
                    name text,
                    cat CHAR(9));"""
        
        #Query per l'inserimento dei dati nella tabella
        insertQuery = """INSERT INTO """+ tablename +"""(
                     _id ,name,cat) VALUES(%s,LOWER(%s),LOWER(%s));"""
        
    elif filename == 'centri':
        #Caricamento del worksheet
        sheet = xl.load_workbook("C:/Users/utente/ICon/Dataset_xlsx/" + filename +".xlsx")
        #sheet = xl.load_workbook("C:/Users/franc/ICon/Dataset_xlsx/" + filename +".xlsx")
        #sheet = xl.load_workbook("C:/Users/nico9/ICon/Dataset_xlsx/" + filename +".xlsx")
        
        table = sheet['cent']
        
        #Query per la creazione della tabella del file centri.xlsx
        queryCreate = """CREATE TABLE """+ tablename + """(
                    id CHAR(12),
                    name VARCHAR(50),
                    lat CHAR(12),
                    longi CHAR(12));"""
        #Query per l'inserimento dei dati nella tabella
        insertQuery = """ INSERT INTO """+ tablename + """(
                     id,name,lat,longi) VALUES(%s,%s,%s,%s);"""
    elif filename == 'centri_cure':
        #Caricamento del worksheet
        sheet = xl.load_workbook("C:/Users/utente/ICon/Dataset_xlsx/" + filename +".xlsx")
        #sheet = xl.load_workbook("C:/Users/franc/ICon/Dataset_xlsx/" + filename +".xlsx")
        #sheet = xl.load_workbook("C:/Users/nico9/ICon/Dataset_xlsx/" + filename +".xlsx")
        
        table = sheet['cure']
        
        #Query per la creazione della tabella del file centri_cure.xlsx
        queryCreate = """CREATE TABLE """+ tablename + """(
                    idCentri CHAR(10),
                    idCure CHAR(10));"""
        #Query per l'inserimento dei dati nella tabella
        insertQuery = """ INSERT INTO """+ tablename + """(
                     idCentri,idCure) VALUES(%s,%s);"""
        
    try:
        cursor.execute(queryDrop)
        cursor.execute(queryCreate)
        conn.commit()
    except sql.ProgrammingError:
        pass
    
    for row in table.rows:
        
        if filename == 'symptoms2':
             _id = row[1].value
             name = row[3].value
             cat = row[9].value
             values = (_id,name,cat)
             #Esecuzione query
             cursor.execute(insertQuery,values)
        elif filename == 'diffsydiw':
            syd = row[0].value
            did = row[1].value
            values = (syd,did)
            #Esecuzione query
            cursor.execute(insertQuery,values)
        elif filename == 'diagn_title':
            id = row[0].value
            title = row[1].value
            cat = row[2].value
            values = (id,title,cat)
            #Esecuzione query
            cursor.execute(insertQuery,values)
        elif filename == 'centri':
            id = row[0].value
            name = row[1].value
            lat = row[2].value
            long = row[3].value
            values = (id,name,lat,long)
            #Esecuzione query
            cursor.execute(insertQuery,values)
        elif filename == 'centri_cure':
            idCentri = row[0].value
            idCure = row[1].value
            values = (idCentri,idCure)
            #Esecuzione query
            cursor.execute(insertQuery,values)
            
            
    cursor.close()      
    #Commit della transazione
    conn.commit()
    

def checkSymWithErr(conn,value):
    """
    Funzione che verifica se il sintomo è presente nel database
    
    Parameters
    ----------
    conn: connection
        Connessione al database
    value: str
        Nome del sintomo dai cui verificare la presenza
        
    Returns
    -------
    boolean
        True se il sintomo è presente, False altrimenti
    """
    cursor = conn.cursor(buffered = True)
    querySym = """SELECT _id from sym where name='""" + value +"""';""" 
    cursor.execute(querySym)
    check = cursor.fetchall()
    if not check:
        return False
    else:
        return True

def checkDiseaseWithErr(conn,value):
    """
    Funzione che verifica la presenza nel database di una malattia e restituisce
    
    Parameters
    ----------
    conn: connection
        Connessione al database
    value: Nome della malattia di cui veridicare la presenza
    
    Returns
    -------
    check: str
         Variabile che contiene l'ID della malattia se essa è presente nel 
         database. In caso contrario è vuota
    """
    cursor = conn.cursor(buffered = True)
    queryDis = """SELECT id from diagn where title like '%"""+value+"""%';"""
    cursor.execute(queryDis)
    check = cursor.fetchall()
    check = check[0][0]
    return check
        
    
def searchDiagn(conn,list):
    """
    Funzione che, data una lista di sintomi, restituisce la lista di tutte le
    diagnosi ad essi corrispondenti
    
    Parameters
    ----------
    conn: connection
        Connessione al database
    list: list
        Lista contenente i nomi dei sintomi
        
    Returns
    -------
    idDiseases: list
        Lista di liste. Ogni lista contiene gli di delle diagnosi relative ad un sintomo
    """
    import numpy as np
    
    cursor = conn.cursor(buffered = True)
    idDiseases = []
    idSymptoms = []
    
    for i in list:
        #Query che preleva gli id corrispondenti ai sintomi
        querySym = """SELECT DISTINCT _id from sym where name='""" + i +"""';""" 
        cursor.execute(querySym)
        idSymptoms.append(cursor.fetchall()) 

    for i in idSymptoms:
        i = np.asarray(i)
        #Query che restituisce l'id delle malattie collegate all'id dei sintomi(colonna did nella tabella)
        queryDid = """SELECT did from diff where syd='"""+ i[0][0] +"""';"""
        cursor.execute(queryDid)
        idDiseases.append(cursor.fetchall())
    
    return idDiseases
    
        
def searchSymCategories(conn,list):
    """
    Funzione che, ad ogni sintomo presente nella lista, associa la rispettiva 
    categoria ci appartenenza
    
    Parameters
    ----------
    conn: connection
        Connessione al database
    list: list
        Lista di sintomi
    
    Returns
    -------
    categories: list
        Lista di categorie
    """
    cursor = conn.cursor(buffered = True)
    categories = []
    
    for i in list:
        #Query che preleva le categorie corrispondenti ai sintomi
        querySym = """SELECT cat from sym where name='""" + i +"""';""" 
        cursor.execute(querySym)
        categories.append(cursor.fetchall())
    return categories
   
    
def findDiagnName(cursor, idDiagn):
    """
    Funzione che restituisce il nome di una diagnosi dato il suo id
    
    Parameters
    ----------
    cursor: cursor
        Cursore per la query
    idDiagn: str
        id della diagnosi di cui si vuole scoprire il nome
    
    Returns
    -------
    cursor: str
        Nome della diagnosi
    """
    queryName = """SELECT title from diagn where id ='"""+idDiagn[0]+"""';"""
    cursor.execute(queryName) 
    cursor = cursor.fetchall()
    cursor = cursor[0][0]
    if "_" in cursor:
        pos = cursor.index("_")
        cursor = cursor[0:pos]
    return cursor

def searchCatDiagn(conn,diagn):
    """
    Funzione che restituisce la categoria a cui una diagnosi appartiene
    
    Parameters
    ----------
    conn: connection
        Connessione al database
    diagn: str
        Nome della diagnosi
    
    Returns
    -------
    cursor: str
        Nome della categoria a cui la diagnosi appartiene
    """
    query = """SELECT cat FROM diagn WHERE id = """+ diagn + """;"""
    cursor = conn.cursor(buffered = True)
    cursor.execute(query)
    cursor = cursor.fetchall()
    return cursor[0]

def findHospitals(conn, diagn):
    """
    Funzione che trova gli ospedali che curano una malattia
    
    Parameters
    ----------
    conn: connection
        Connessione al database
    diagn: str
        Nome della diagnosi
    
    Returns
    -------
    cursor: list
        Lista di triple (nome-latitudine-longitudine), ognuna rappresentante un
        ospedale che cura la malattia
    """
    cursor = conn.cursor(buffered = True)
    query = """SELECT cent.name, cent.lat, cent.longi FROM cent INNER JOIN cent_cure on cent.id = cent_cure.idCentri WHERE idCure ='"""+ diagn +"""';"""
    cursor.execute(query)
    cursor = cursor.fetchall()
    return cursor

def closeConn(conn):
    """
    Funzione che chiude la connessione al database
    
    Parameters
    ----------
    conn: connection
        Connessione al database
    """
    conn.close()
    
    